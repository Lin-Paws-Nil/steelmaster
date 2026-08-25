"""
BBS Export Service

Generates formatted Excel spreadsheets from calculated Bar Bending Schedule data.
Produces two sheets:
    - "BBS Details": Full line-by-line breakdown with engineering headers
    - "Weight Summary": Aggregation by bar diameter for procurement
"""

import io
from typing import List, Optional, Union

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from backend.app.models.bbs import BBSRow


# Column rename mapping: internal field -> engineering display header
COLUMN_HEADERS = {
    "beam_id": "Beam ID",
    "bar_type": "Bar Type",
    "diameter": "Dia (mm)",
    "count": "No. of Bars",
    "shape_code": "Shape",
    "cutting_length_m": "Cut Length (m)",
    "total_weight_kg": "Total Weight (kg)",
}

SUMMARY_HEADERS = {
    "diameter": "Dia (mm)",
    "total_weight_kg": "Total Weight (kg)",
    "bar_count_total": "Total No. of Bars",
}


class BBSExporter:
    """Exports BBS data to formatted Excel spreadsheets."""

    def generate_excel(
        self,
        bbs_data: List[BBSRow],
        output_path: Optional[str] = None,
    ) -> Union[str, bytes]:
        """
        Generate a formatted Excel workbook from BBS data.

        Args:
            bbs_data: List of BBSRow Pydantic models from the calculation engine.
            output_path: File path to write the .xlsx file.
                         If None, returns the file as bytes (for in-memory streaming).

        Returns:
            If output_path is provided: the output path string.
            If output_path is None: bytes of the Excel file.

        Raises:
            ValueError: If bbs_data is empty.
        """
        if not bbs_data:
            raise ValueError("Cannot generate BBS report from empty data.")

        # Convert to DataFrame
        df = pd.DataFrame([row.model_dump() for row in bbs_data])
        df_display = df.rename(columns=COLUMN_HEADERS)

        # Create summary aggregation by diameter
        summary = (
            df.groupby("diameter")
            .agg(
                total_weight_kg=("total_weight_kg", "sum"),
                bar_count_total=("count", "sum"),
            )
            .reset_index()
            .sort_values("diameter")
        )
        summary_display = summary.rename(columns=SUMMARY_HEADERS)

        # Add totals row to summary
        totals = pd.DataFrame([{
            "Dia (mm)": "TOTAL",
            "Total Weight (kg)": summary["total_weight_kg"].sum(),
            "Total No. of Bars": summary["bar_count_total"].sum(),
        }])
        summary_display = pd.concat([summary_display, totals], ignore_index=True)

        # Write to Excel
        if output_path is not None:
            buffer = output_path
        else:
            buffer = io.BytesIO()

        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df_display.to_excel(writer, sheet_name="BBS Details", index=False)
            summary_display.to_excel(writer, sheet_name="Weight Summary", index=False)

        # Apply formatting
        if output_path is not None:
            self._apply_formatting(output_path)
            return output_path
        else:
            buffer.seek(0)
            raw_bytes = buffer.read()
            # Apply formatting by reloading from bytes
            formatted = self._apply_formatting_bytes(raw_bytes)
            return formatted

    def generate_excel_bytes(self, bbs_data: List[BBSRow]) -> bytes:
        """
        Generate Excel file and return as bytes (for API streaming).

        Args:
            bbs_data: List of BBSRow objects.

        Returns:
            Bytes of the formatted .xlsx file.
        """
        result = self.generate_excel(bbs_data, output_path=None)
        if isinstance(result, bytes):
            return result
        raise RuntimeError("Expected bytes output")

    def _apply_formatting(self, filepath: str) -> None:
        """Apply formatting to an on-disk Excel file."""
        wb = load_workbook(filepath)
        self._format_workbook(wb)
        wb.save(filepath)

    def _apply_formatting_bytes(self, raw_bytes: bytes) -> bytes:
        """Apply formatting to Excel bytes and return formatted bytes."""
        buffer_in = io.BytesIO(raw_bytes)
        wb = load_workbook(buffer_in)
        self._format_workbook(wb)
        buffer_out = io.BytesIO()
        wb.save(buffer_out)
        buffer_out.seek(0)
        return buffer_out.read()

    def _format_workbook(self, wb) -> None:
        """Apply formatting to all sheets in the workbook."""
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_text_font = Font(bold=True, color="FFFFFF", size=11)
        thin_border = Border(
            left=Side("thin"),
            right=Side("thin"),
            top=Side("thin"),
            bottom=Side("thin"),
        )
        weight_format = "0.00"    # 2 decimal places for weights (kg)
        length_format = "0.000"   # 3 decimal places for cutting lengths (m)

        for ws in wb.worksheets:
            # Format header row
            for cell in ws[1]:
                cell.font = header_text_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            # Format data rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")
                    if isinstance(cell.value, float):
                        header_val = ws.cell(1, cell.column).value or ""
                        if "Weight" in header_val:
                            cell.number_format = weight_format
                        elif "Length" in header_val or "Cut" in header_val:
                            cell.number_format = length_format
                        else:
                            cell.number_format = weight_format

            # Auto-adjust column widths
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_width = 0
                for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            cell_len = len(str(cell.value))
                            max_width = max(max_width, cell_len)
                ws.column_dimensions[col_letter].width = min(max_width + 4, 30)

            # Bold the totals row in Weight Summary
            if ws.title == "Weight Summary":
                last_row = ws.max_row
                for cell in ws[last_row]:
                    cell.font = Font(bold=True, size=11)
