"""Estimate router - handles steel estimation from parsed elements."""

import csv
import io

from fastapi import APIRouter, Body
from fastapi.responses import StreamingResponse

from backend.app.models.schemas import StructuralElement, ProjectEstimate
from backend.app.services.steel_estimator import estimate_project, estimate_element

router = APIRouter()


@router.post("/estimate", response_model=ProjectEstimate)
async def create_estimate(project_name: str, elements: list[StructuralElement]):
    """Estimate steel for a list of structural elements."""
    return estimate_project(project_name, elements)


@router.post("/estimate-single")
async def estimate_single_element(element: StructuralElement):
    """Estimate steel for a single element (useful for quick calculations)."""
    return estimate_element(element)


@router.post("/download-csv")
async def download_csv(estimate: dict = Body(...)):
    """Generate a downloadable CSV file from estimation results."""
    output = io.StringIO()
    writer = csv.writer(output)

    # Header
    project_name = estimate.get("project_name", "Steel Estimate")
    writer.writerow(["PROJECT STEEL ESTIMATION REPORT"])
    writer.writerow(["Project:", project_name])
    writer.writerow(["Total Steel (kg):", estimate.get("total_steel_kg", 0)])
    writer.writerow(["Total Steel (tonnes):", estimate.get("total_steel_tons", 0)])
    writer.writerow([])

    # Summary by type
    writer.writerow(["SUMMARY BY ELEMENT TYPE"])
    writer.writerow(["Element Type", "Weight (kg)"])
    for etype, weight in estimate.get("summary_by_type", {}).items():
        writer.writerow([etype, round(weight, 2)])
    writer.writerow([])

    # Summary by diameter
    writer.writerow(["SUMMARY BY BAR DIAMETER"])
    writer.writerow(["Diameter", "Weight (kg)"])
    for dia, weight in estimate.get("summary_by_diameter", {}).items():
        writer.writerow([dia, round(weight, 2)])
    writer.writerow([])

    # Detailed breakdown
    writer.writerow(["DETAILED BREAKDOWN"])
    writer.writerow([
        "Element Label", "Element Type", "Width (mm)", "Depth (mm)",
        "Span/Length (mm)", "Bar Type", "Diameter (mm)", "Count",
        "Length (m)", "Wt/m (kg/m)", "Total Weight (kg)"
    ])

    for el in estimate.get("elements", []):
        element = el.get("element", {})
        rebars = el.get("rebars", [])
        for bar in rebars:
            writer.writerow([
                element.get("label", ""),
                element.get("element_type", ""),
                element.get("width", ""),
                element.get("depth", ""),
                element.get("length", ""),
                bar.get("bar_type", ""),
                bar.get("diameter", ""),
                bar.get("count", ""),
                round(bar.get("length", 0), 2),
                round(bar.get("weight_per_meter", 0), 3),
                round(bar.get("total_weight", 0), 2),
            ])

    writer.writerow([])
    writer.writerow(["", "", "", "", "", "", "", "", "", "TOTAL:", round(estimate.get("total_steel_kg", 0), 2)])

    output.seek(0)
    filename = f"{project_name.replace(' ', '_')}_steel_estimate.csv"

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/download-excel")
async def download_excel(estimate: dict = Body(...)):
    """Generate a downloadable Excel file from estimation results."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(status_code=500, detail="openpyxl not installed. Falling back to CSV.")

    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"

    project_name = estimate.get("project_name", "Steel Estimate")
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=11)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_text = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    ws["A1"] = "STEEL ESTIMATION REPORT"
    ws["A1"].font = header_font
    ws["A3"] = "Project:"
    ws["B3"] = project_name
    ws["B3"].font = Font(bold=True)
    ws["A4"] = "Total Steel (kg):"
    ws["B4"] = estimate.get("total_steel_kg", 0)
    ws["B4"].font = Font(bold=True, size=12)
    ws["A5"] = "Total Steel (tonnes):"
    ws["B5"] = estimate.get("total_steel_tons", 0)

    # Summary by type
    row = 7
    ws.cell(row, 1, "BY ELEMENT TYPE").font = section_font
    row += 1
    for col, header in enumerate(["Element Type", "Weight (kg)"], 1):
        cell = ws.cell(row, col, header)
        cell.font = header_text
        cell.fill = header_fill
    row += 1
    for etype, weight in estimate.get("summary_by_type", {}).items():
        ws.cell(row, 1, etype)
        ws.cell(row, 2, round(weight, 2))
        row += 1

    # Summary by diameter
    row += 1
    ws.cell(row, 1, "BY BAR DIAMETER").font = section_font
    row += 1
    for col, header in enumerate(["Diameter", "Weight (kg)"], 1):
        cell = ws.cell(row, col, header)
        cell.font = header_text
        cell.fill = header_fill
    row += 1
    for dia, weight in estimate.get("summary_by_diameter", {}).items():
        ws.cell(row, 1, dia)
        ws.cell(row, 2, round(weight, 2))
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15

    # --- Sheet 2: Detailed Breakdown ---
    ws2 = wb.create_sheet("Detailed Breakdown")

    headers = [
        "Element Label", "Element Type", "Width (mm)", "Depth (mm)",
        "Span (mm)", "Bar Type", "Dia (mm)", "Count",
        "Length (m)", "Wt/m (kg/m)", "Weight (kg)"
    ]
    for col, header in enumerate(headers, 1):
        cell = ws2.cell(1, col, header)
        cell.font = header_text
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2
    for el in estimate.get("elements", []):
        element = el.get("element", {})
        rebars = el.get("rebars", [])
        for bar in rebars:
            ws2.cell(row, 1, element.get("label", ""))
            ws2.cell(row, 2, element.get("element_type", ""))
            ws2.cell(row, 3, element.get("width", 0))
            ws2.cell(row, 4, element.get("depth", 0))
            ws2.cell(row, 5, element.get("length", 0))
            ws2.cell(row, 6, bar.get("bar_type", ""))
            ws2.cell(row, 7, bar.get("diameter", 0))
            ws2.cell(row, 8, bar.get("count", 0))
            ws2.cell(row, 9, round(bar.get("length", 0), 2))
            ws2.cell(row, 10, round(bar.get("weight_per_meter", 0), 3))
            ws2.cell(row, 11, round(bar.get("total_weight", 0), 2))
            row += 1

    # Total row
    ws2.cell(row, 10, "TOTAL:")
    ws2.cell(row, 10).font = Font(bold=True)
    ws2.cell(row, 11, round(estimate.get("total_steel_kg", 0), 2))
    ws2.cell(row, 11).font = Font(bold=True)

    for col in range(1, 12):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 14

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{project_name.replace(' ', '_')}_steel_estimate.xlsx"

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
